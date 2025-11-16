import os
import argparse
import logging
from tqdm import tqdm

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.backends.cudnn as cudnn
from monai.networks.nets import SwinUNETR, UNETR

from openpyxl import Workbook, load_workbook

from picai_dataset import get_sdcl_dataloaders
from models.unet3d import UNet3D
from models.resunet3d import ResUNet3D
from losses import (
    DiceCELoss,
    mix_loss,
    mix_mse_loss,
    mix_max_kl_loss,
    get_XOR_region,
)


def generate_3d_mask(img):
    """
    img: [B, C, H, W, D]
    Returns:
      img_mask:  [B, 1, H, W, D]  (for mixing input images)
      loss_mask: [B, H, W, D]     (for region-weighted losses)
    """
    b, c, h, w, d = img.shape
    device = img.device

    mask = torch.ones((b, 1, h, w, d), device=device)
    loss_mask = torch.ones((b, h, w, d), device=device)

    patch_h = int(h * 2 / 3)
    patch_w = int(w * 2 / 3)
    patch_d = int(max(1, d * 2 / 3))

    top = np.random.randint(0, max(1, h - patch_h + 1))
    left = np.random.randint(0, max(1, w - patch_w + 1))
    front = np.random.randint(0, max(1, d - patch_d + 1))

    mask[:, :, top: top + patch_h, left: left + patch_w, front: front + patch_d] = 0.0
    loss_mask[:, top: top + patch_h, left: left + patch_w, front: front + patch_d] = 0.0

    return mask, loss_mask



class BCPNet(nn.Module):
    """Wrapper for student / teacher networks.

    Supports:
      - "unet"     -> custom UNet3D
      - "resunet"  -> custom ResUNet3D
      - "swinunetr"-> MONAI SwinUNETR
      - "unetr"    -> MONAI UNETR
    """
    def __init__(self, model_name: str, in_chns: int, num_classes: int, ema: bool = False):
        super().__init__()
        model_name = model_name.lower()
        if model_name == "unet":
            self.net = UNet3D(in_channels=in_chns, num_classes=num_classes)
        elif model_name == "resunet":
            self.net = ResUNet3D(in_channels=in_chns, num_classes=num_classes)
        elif model_name == "swinunetr":
            # image size is (H, W, D) = (144, 128, 16) for this project
            self.net = SwinUNETR(
                img_size=(144, 128, 16),
                in_channels=in_chns,
                out_channels=num_classes,
                feature_size=48,
                use_checkpoint=True,
            )
        elif model_name == "unetr":
            self.net = UNETR(
                img_size=(144, 128, 16),
                in_channels=in_chns,
                out_channels=num_classes,
                feature_size=16,
                hidden_size=768,
                mlp_dim=3072,
                num_heads=12,
                pos_embed="perceptron",
                norm_name="instance",
                res_block=True,
                dropout_rate=0.0,
            )
        else:
            raise ValueError(f"Unknown model_name {model_name}")

        # for EMA teachers we don't want gradients flowing
        if ema:
            for p in self.net.parameters():
                p.detach_()

    def forward(self, x):
        return self.net(x)



def update_ema_variables(model, ema_model, alpha=0.99):
    """
    EMA update for teacher model.
    Updates both parameters and buffers (e.g., norm running stats).
    """
    # EMA for parameters
    for ema_param, param in zip(ema_model.parameters(), model.parameters()):
        ema_param.data.mul_(alpha).add_(param.data, alpha=1 - alpha)

    # keep buffers (like running_mean / running_var) in sync
    for ema_buf, buf in zip(ema_model.buffers(), model.buffers()):
        ema_buf.data.copy_(buf.data)
def dice_metric_from_logits(logits, targets, num_classes: int):
    """
    Compute mean foreground Dice for evaluation.
    """
    with torch.no_grad():
        probs = F.softmax(logits, dim=1)
        preds = torch.argmax(probs, dim=1)

        dices = []
        for cls in range(1, num_classes):  # ignore background = 0
            pred_c = (preds == cls).float()
            target_c = (targets == cls).float()
            inter = (pred_c * target_c).sum()
            denom = pred_c.sum() + target_c.sum()
            if denom > 0:
                dices.append((2.0 * inter + 1e-5) / (denom + 1e-5))
        if len(dices) == 0:
            return 0.0
        return float(torch.mean(torch.stack(dices)))



def get_parser():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--root_path",
        type=str,
        required=True,
        help="Root path of PiCAI data (train/ valid/ unlab_data/)",
    )
    parser.add_argument("--gpu", type=str, default="0")
    parser.add_argument("--batch_size", type=int, default=2)
    parser.add_argument("--max_epoch", type=int, default=300)
    parser.add_argument("--num_classes", type=int, default=2)
    parser.add_argument("--u_weight", type=float, default=0.5)
    parser.add_argument("--seed", type=int, default=1337)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--snapshot_path", type=str, default="./runs_sdcl_picai")

    parser.add_argument(
        "--model1",
        type=str,
        default="unet",
        help="Backbone for student 1 (options: unet, resunet, swinunetr, unetr)",
    )
    parser.add_argument(
        "--model2",
        type=str,
        default="resunet",
        help="Backbone for student 2 (options: unet, resunet, swinunetr, unetr)",
    )

    return parser


def train(args):
    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

    snapshot_path = args.snapshot_path
    os.makedirs(snapshot_path, exist_ok=True)

    # logging
    logging.basicConfig(
        filename=os.path.join(snapshot_path, "log.txt"),
        level=logging.INFO,
        format="[%(asctime)s.%(msecs)03d] %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.getLogger().addHandler(logging.StreamHandler())
    logging.info(str(args))

    # Excel metrics file
    excel_path = os.path.join(snapshot_path, "metrics.xlsx")
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrics"
        ws.append(["Epoch", "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA"])
        wb.save(excel_path)

    # reproducibility (without forcing deterministic ops that break MaxPool3d)
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    torch.cuda.manual_seed_all(args.seed)
    cudnn.benchmark = False
    # DO NOT call torch.use_deterministic_algorithms(True) here

    # dataloaders
    (
        lab_loader_a,
        lab_loader_b,
        unlab_loader_a,
        unlab_loader_b,
        val_loader,
    ) = get_sdcl_dataloaders(
        args.root_path,
        batch_size=args.batch_size,
        num_workers=4,
        seed=args.seed,
    )

    num_classes = args.num_classes
    in_chns = 3

    # models
    model = BCPNet(model_name=args.model1, in_chns=in_chns, num_classes=num_classes).cuda()
    model2 = BCPNet(model_name=args.model2, in_chns=in_chns, num_classes=num_classes).cuda()
    # EMA teacher mirrors student 1's architecture
    ema_model = BCPNet(model_name=args.model1, in_chns=in_chns, num_classes=num_classes, ema=True).cuda()

    optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)
    optimizer2 = torch.optim.Adam(model2.parameters(), lr=args.lr)

    # you can use this for fully supervised warmup if needed (currently unused)
    criterion_sup = DiceCELoss(num_classes=num_classes)

    max_epoch = args.max_epoch
    best_dice_ema = 0.0      # best EMA dice
    best_dice_student1 = 0.0 # best student1 dice
    global_step = 0

    for epoch in range(max_epoch):
        model.train()
        model2.train()
        ema_model.train()

        loader_iter = zip(lab_loader_a, lab_loader_b, unlab_loader_a, unlab_loader_b)
        progress = tqdm(loader_iter, ncols=100, desc=f"Epoch {epoch + 1}/{max_epoch}")

        for (img_a, lab_a), (img_b, lab_b), (unimg_a, _), (unimg_b, _) in progress:
            img_a = img_a.cuda()
            lab_a = lab_a.cuda().long().squeeze(1)  # [B, H, W, D]
            img_b = img_b.cuda()
            lab_b = lab_b.cuda().long().squeeze(1)
            unimg_a = unimg_a.cuda()
            unimg_b = unimg_b.cuda()

            # teacher pseudo labels for unlabeled images
            with torch.no_grad():
                pre_a = ema_model(unimg_a)
                pre_b = ema_model(unimg_b)

                plab_a = torch.argmax(F.softmax(pre_a, dim=1), dim=1)
                plab_b = torch.argmax(F.softmax(pre_b, dim=1), dim=1)

                img_mask, loss_mask = generate_3d_mask(img_a)
                # loss_mask: [B, H, W, D], used directly in losses

            # mix labeled & unlabeled (BCP style)
            net_input_l = unimg_a * img_mask + img_b * (1.0 - img_mask)
            net_input_unl = img_a * img_mask + unimg_b * (1.0 - img_mask)

            out_l = model(net_input_l)
            out_unl = model(net_input_unl)

            out_l_2 = model2(net_input_l)
            out_unl_2 = model2(net_input_unl)

            # SDCL mix losses for net1
            l_dice, l_ce = mix_loss(
                out_l,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
            )
            unl_dice, unl_ce = mix_loss(
                out_unl,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
            )

            # SDCL mix losses for net2
            l_dice_2, l_ce_2 = mix_loss(
                out_l_2,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
            )
            unl_dice_2, unl_ce_2 = mix_loss(
                out_unl_2,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
            )

            loss_ce = unl_ce + l_ce
            loss_dice = unl_dice + l_dice

            loss_ce_2 = unl_ce_2 + l_ce_2
            loss_dice_2 = unl_dice_2 + l_dice_2

            with torch.no_grad():
                diff_mask1 = get_XOR_region(out_l, out_l_2)
                diff_mask2 = get_XOR_region(out_unl, out_unl_2)

            # discrepancy-aware MSE & KL (net1)
            net1_mse_loss_lab = mix_mse_loss(
                out_l,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net1_mse_loss_unlab = mix_mse_loss(
                out_unl,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )
            net1_kl_loss_lab = mix_max_kl_loss(
                out_l,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net1_kl_loss_unlab = mix_max_kl_loss(
                out_unl,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )

            # discrepancy-aware MSE & KL (net2)
            net2_mse_loss_lab = mix_mse_loss(
                out_l_2,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net2_mse_loss_unlab = mix_mse_loss(
                out_unl_2,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )
            net2_kl_loss_lab = mix_max_kl_loss(
                out_l_2,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net2_kl_loss_unlab = mix_max_kl_loss(
                out_unl_2,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )

            loss1 = (
                (loss_dice + loss_ce) / 2.0
                + 0.5 * (net1_mse_loss_lab + net1_mse_loss_unlab)
                + 0.05 * (net1_kl_loss_lab + net1_kl_loss_unlab)
            )
            loss2 = (
                (loss_dice_2 + loss_ce_2) / 2.0
                + 0.5 * (net2_mse_loss_lab + net2_mse_loss_unlab)
                + 0.05 * (net2_kl_loss_lab + net2_kl_loss_unlab)
            )

            optimizer.zero_grad()
            loss1.backward()
            optimizer.step()

            optimizer2.zero_grad()
            loss2.backward()
            optimizer2.step()

            update_ema_variables(model, ema_model, alpha=0.99)

            global_step += 1
            progress.set_postfix(
                {
                    "loss1": f"{loss1.item():.4f}",
                    "loss2": f"{loss2.item():.4f}",
                }
            )

        # =====================
        # Validation
        # =====================
        model.eval()
        model2.eval()
        ema_model.eval()

        dices1 = []
        dices2 = []
        dices_ema = []
        with torch.no_grad():
            for img_v, lab_v in val_loader:
                img_v = img_v.cuda()
                lab_v = lab_v.cuda().long().squeeze(1)

                out1 = model(img_v)
                out2 = model2(img_v)
                out_ema = ema_model(img_v)

                dices1.append(dice_metric_from_logits(out1, lab_v, num_classes))
                dices2.append(dice_metric_from_logits(out2, lab_v, num_classes))
                dices_ema.append(dice_metric_from_logits(out_ema, lab_v, num_classes))

        mean_dice1 = float(np.mean(dices1))
        mean_dice2 = float(np.mean(dices2))
        mean_dice_ema = float(np.mean(dices_ema))

        logging.info(
            f"Epoch {epoch + 1}: dice1={mean_dice1:.4f}, dice2={mean_dice2:.4f}, dice_ema={mean_dice_ema:.4f}"
        )

        # Save best EMA model
        if mean_dice_ema > best_dice_ema:
            best_dice_ema = mean_dice_ema
            torch.save(
                ema_model.state_dict(),
                os.path.join(snapshot_path, "best_ema_model.pth"),
            )
            logging.info(f"*** Updated best EMA model: dice={best_dice_ema:.4f}")

        # Save best student1 model
        if mean_dice1 > best_dice_student1:
            best_dice_student1 = mean_dice1
            torch.save(
                model.state_dict(),
                os.path.join(snapshot_path, "best_student1_model.pth"),
            )

        # Append metrics to Excel
        wb = load_workbook(excel_path)
        ws = wb["Metrics"]
        ws.append(
            [epoch + 1, mean_dice1, mean_dice2, mean_dice_ema, best_dice_ema]
        )
        wb.save(excel_path)

    logging.info(f"Training finished. Best EMA Dice: {best_dice_ema:.4f}")


if __name__ == "__main__":
    parser = get_parser()
    args = parser.parse_args()
    train(args)




# import os
# import argparse
# import logging
# from tqdm import tqdm

# import numpy as np
# import torch
# import torch.nn as nn
# import torch.nn.functional as F
# import torch.backends.cudnn as cudnn
from monai.networks.nets import SwinUNETR, UNETR

# from openpyxl import Workbook, load_workbook
# import os

# from picai_dataset import get_sdcl_dataloaders
# from models.unet3d import UNet3D
# from models.resunet3d import ResUNet3D
# from losses import (
#     DiceCELoss,
#     mix_loss,
#     mix_mse_loss,
#     mix_max_kl_loss,
#     get_XOR_region,
# )


# def generate_3d_mask(img):
#     """
#     img: [B, C, H, W, D]
#     Returns:
#       img_mask: [B, 1, H, W, D]
#       loss_mask: [B, H, W, D]
#     """
#     b, c, h, w, d = img.shape
#     device = img.device

#     mask = torch.ones((b, 1, h, w, d), device=device)
#     loss_mask = torch.ones((b, h, w, d), device=device)

#     patch_h = int(h * 2 / 3)
#     patch_w = int(w * 2 / 3)
#     patch_d = int(max(1, d * 2 / 3))

#     top = np.random.randint(0, max(1, h - patch_h + 1))
#     left = np.random.randint(0, max(1, w - patch_w + 1))
#     front = np.random.randint(0, max(1, d - patch_d + 1))

#     mask[:, :, top : top + patch_h, left : left + patch_w, front : front + patch_d] = 0.0
#     loss_mask[:, top : top + patch_h, left : left + patch_w, front : front + patch_d] = 0.0

#     return mask, loss_mask


# class BCPNet(nn.Module):
#     """
#     Wrapper for student / teacher networks.
#     """
#     def __init__(self, model_name: str, in_chns: int, num_classes: int, ema: bool = False):
#         super().__init__()
#         model_name = model_name.lower()
#         if model_name == "unet":
#             self.net = UNet3D(in_channels=in_chns, num_classes=num_classes)
#         elif model_name == "resunet":
#             self.net = ResUNet3D(in_channels=in_chns, num_classes=num_classes)
#         else:
#             raise ValueError(f"Unknown model_name {model_name}")

#         if ema:
#             for p in self.net.parameters():
#                 p.detach_()

#     def forward(self, x):
#         return self.net(x)


# def update_ema_variables(model, ema_model, alpha=0.99):
#     """
#     EMA update for teacher model.
#     """
#     for ema_param, param in zip(ema_model.parameters(), model.parameters()):
#         ema_param.data.mul_(alpha).add_(param.data, alpha=1 - alpha)


# def dice_metric_from_logits(logits, targets, num_classes: int):
#     """
#     Compute mean foreground Dice for evaluation.
#     """
#     with torch.no_grad():
#         probs = F.softmax(logits, dim=1)
#         preds = torch.argmax(probs, dim=1)

#         dices = []
#         for cls in range(1, num_classes):  # ignore background
#             pred_c = (preds == cls).float()
#             target_c = (targets == cls).float()
#             inter = (pred_c * target_c).sum()
#             denom = pred_c.sum() + target_c.sum()
#             if denom > 0:
#                 dices.append((2.0 * inter + 1e-5) / (denom + 1e-5))
#         if len(dices) == 0:
#             return 0.0
#         return float(torch.mean(torch.stack(dices)))


# def get_parser():
#     parser = argparse.ArgumentParser()

#     parser.add_argument("--root_path", type=str, required=True,
#                         help="Root path of PiCAI data (train/ valid/ unlab_data/)")
#     parser.add_argument("--gpu", type=str, default="0")
#     parser.add_argument("--batch_size", type=int, default=2)
#     parser.add_argument("--max_epoch", type=int, default=300)
#     parser.add_argument("--num_classes", type=int, default=2)
#     parser.add_argument("--u_weight", type=float, default=0.5)
#     parser.add_argument("--seed", type=int, default=1337)
#     parser.add_argument("--lr", type=float, default=1e-3)
#     parser.add_argument("--snapshot_path", type=str, default="./runs_sdcl_picai")

#     return parser


# def train(args):
#     os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

#     snapshot_path = args.snapshot_path
#     os.makedirs(snapshot_path, exist_ok=True)

#     logging.basicConfig(
#         filename=os.path.join(snapshot_path, "log.txt"),
#         level=logging.INFO,
#         format="[%(asctime)s.%(msecs)03d] %(message)s",
#         datefmt="%H:%M:%S",
#     )
#     logging.getLogger().addHandler(logging.StreamHandler())
#     logging.info(str(args))

#     # reproducibility
#     np.random.seed(args.seed)
#     torch.manual_seed(args.seed)
#     torch.cuda.manual_seed_all(args.seed)
#     cudnn.benchmark = False
#     # torch.use_deterministic_algorithms(True)

#     # dataloaders
#     (
#         lab_loader_a,
#         lab_loader_b,
#         unlab_loader_a,
#         unlab_loader_b,
#         val_loader,
#     ) = get_sdcl_dataloaders(
#         args.root_path,
#         batch_size=args.batch_size,
#         num_workers=4,
#         seed=args.seed,
#     )

#     num_classes = args.num_classes
#     in_chns = 3

#     # models
#     model = BCPNet(model_name="unet", in_chns=in_chns, num_classes=num_classes).cuda()
#     model2 = BCPNet(model_name="resunet", in_chns=in_chns, num_classes=num_classes).cuda()
#     ema_model = BCPNet(model_name="unet", in_chns=in_chns, num_classes=num_classes, ema=True).cuda()

#     optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)
#     optimizer2 = torch.optim.Adam(model2.parameters(), lr=args.lr)

#     criterion_sup = DiceCELoss(num_classes=num_classes)

#     max_epoch = args.max_epoch
#     best_dice = 0.0
#     best_dice2 = 0.0
#     global_step = 0

#     for epoch in range(max_epoch):
#         model.train()
#         model2.train()
#         ema_model.train()

#         loader_iter = zip(lab_loader_a, lab_loader_b, unlab_loader_a, unlab_loader_b)
#         progress = tqdm(loader_iter, ncols=100, desc=f"Epoch {epoch+1}/{max_epoch}")

#         for (img_a, lab_a), (img_b, lab_b), (unimg_a, _), (unimg_b, _) in progress:
#             img_a = img_a.cuda()
#             lab_a = lab_a.cuda().long().squeeze(1)
#             img_b = img_b.cuda()
#             lab_b = lab_b.cuda().long().squeeze(1)
#             unimg_a = unimg_a.cuda()
#             unimg_b = unimg_b.cuda()

#             with torch.no_grad():
#                 pre_a = ema_model(unimg_a)
#                 pre_b = ema_model(unimg_b)

#                 plab_a = torch.argmax(F.softmax(pre_a, dim=1), dim=1)
#                 plab_b = torch.argmax(F.softmax(pre_b, dim=1), dim=1)

#                 img_mask, loss_mask = generate_3d_mask(img_a)
#                 # broadcast loss_mask to shape [B, H, W, D]
#                 loss_mask = loss_mask

#             # mix labeled & unlabeled
#             net_input_l = unimg_a * img_mask + img_b * (1.0 - img_mask)
#             net_input_unl = img_a * img_mask + unimg_b * (1.0 - img_mask)

#             out_l = model(net_input_l)
#             out_unl = model(net_input_unl)

#             out_l_2 = model2(net_input_l)
#             out_unl_2 = model2(net_input_unl)

#             # SDCL mix losses for net1
#             l_dice, l_ce = mix_loss(
#                 out_l,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#             )
#             unl_dice, unl_ce = mix_loss(
#                 out_unl,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#             )

#             l_dice_2, l_ce_2 = mix_loss(
#                 out_l_2,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#             )
#             unl_dice_2, unl_ce_2 = mix_loss(
#                 out_unl_2,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#             )

#             loss_ce = unl_ce + l_ce
#             loss_dice = unl_dice + l_dice

#             loss_ce_2 = unl_ce_2 + l_ce_2
#             loss_dice_2 = unl_dice_2 + l_dice_2

#             with torch.no_grad():
#                 diff_mask1 = get_XOR_region(out_l, out_l_2)
#                 diff_mask2 = get_XOR_region(out_unl, out_unl_2)

#             net1_mse_loss_lab = mix_mse_loss(
#                 out_l,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net1_mse_loss_unlab = mix_mse_loss(
#                 out_unl,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )
#             net1_kl_loss_lab = mix_max_kl_loss(
#                 out_l,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net1_kl_loss_unlab = mix_max_kl_loss(
#                 out_unl,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )

#             net2_mse_loss_lab = mix_mse_loss(
#                 out_l_2,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net2_mse_loss_unlab = mix_mse_loss(
#                 out_unl_2,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )
#             net2_kl_loss_lab = mix_max_kl_loss(
#                 out_l_2,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net2_kl_loss_unlab = mix_max_kl_loss(
#                 out_unl_2,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )

#             loss1 = (
#                 (loss_dice + loss_ce) / 2.0
#                 + 0.5 * (net1_mse_loss_lab + net1_mse_loss_unlab)
#                 + 0.05 * (net1_kl_loss_lab + net1_kl_loss_unlab)
#             )
#             loss2 = (
#                 (loss_dice_2 + loss_ce_2) / 2.0
#                 + 0.5 * (net2_mse_loss_lab + net2_mse_loss_unlab)
#                 + 0.05 * (net2_kl_loss_lab + net2_kl_loss_unlab)
#             )

#             optimizer.zero_grad()
#             loss1.backward()
#             optimizer.step()

#             optimizer2.zero_grad()
#             loss2.backward()
#             optimizer2.step()

#             update_ema_variables(model, ema_model, alpha=0.99)

#             global_step += 1
#             progress.set_postfix(
#                 {
#                     "loss1": f"{loss1.item():.4f}",
#                     "loss2": f"{loss2.item():.4f}",
#                 }
#             )

#         # Validation
#         model.eval()
#         model2.eval()
#         ema_model.eval()
#         dices1 = []
#         dices2 = []
#         dices_ema = []
#         with torch.no_grad():
#             for img_v, lab_v in val_loader:
#                 img_v = img_v.cuda()
#                 lab_v = lab_v.cuda().long().squeeze(1)

#                 out1 = model(img_v)
#                 out2 = model2(img_v)
#                 out_ema = ema_model(img_v)

#                 dices1.append(dice_metric_from_logits(out1, lab_v, num_classes))
#                 dices2.append(dice_metric_from_logits(out2, lab_v, num_classes))
#                 dices_ema.append(dice_metric_from_logits(out_ema, lab_v, num_classes))

#         mean_dice1 = float(np.mean(dices1))
#         mean_dice2 = float(np.mean(dices2))
#         mean_dice_ema = float(np.mean(dices_ema))

#         logging.info(
#             f"Epoch {epoch+1}: dice1={mean_dice1:.4f}, dice2={mean_dice2:.4f}, dice_ema={mean_dice_ema:.4f}"
#         )

#         if mean_dice_ema > best_dice:
#             best_dice = mean_dice_ema
#             torch.save(
#                 ema_model.state_dict(),
#                 os.path.join(snapshot_path, "best_ema_model.pth"),
#             )
#             logging.info(f"*** Updated best EMA model: dice={best_dice:.4f}")

#         if mean_dice1 > best_dice2:
#             best_dice2 = mean_dice1
#             torch.save(
#                 model.state_dict(),
#                 os.path.join(snapshot_path, "best_student1_model.pth"),
#             )

#     logging.info(f"Training finished. Best EMA Dice: {best_dice:.4f}")


# if __name__ == "__main__":
#     parser = get_parser()
#     args = parser.parse_args()
#     train(args)
