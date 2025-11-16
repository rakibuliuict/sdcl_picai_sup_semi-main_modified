import os
import argparse
import logging
from tqdm import tqdm

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.backends.cudnn as cudnn

from openpyxl import Workbook, load_workbook
from scipy.ndimage import distance_transform_edt, binary_erosion

from picai_dataset import get_sdcl_dataloaders
from models.unet3d import UNet3D
from models.resunet3d import ResUNet3D
from losses import (
    DiceCELoss,
    mix_loss,
    mix_mse_loss,
    mix_max_kl_loss,
    get_XOR_region,
)


# -------------------------
# Utility: BCP mask
# -------------------------
def generate_3d_mask(img):
    """
    img: [B, C, H, W, D]
    Returns:
      img_mask:  [B, 1, H, W, D]  (for mixing input images)
      loss_mask: [B, H, W, D]     (for region-weighted losses)
    """
    b, c, h, w, d = img.shape
    device = img.device

    mask = torch.ones((b, 1, h, w, d), device=device)
    loss_mask = torch.ones((b, h, w, d), device=device)

    patch_h = int(h * 2 / 3)
    patch_w = int(w * 2 / 3)
    patch_d = int(max(1, d * 2 / 3))

    top = np.random.randint(0, max(1, h - patch_h + 1))
    left = np.random.randint(0, max(1, w - patch_w + 1))
    front = np.random.randint(0, max(1, d - patch_d + 1))

    mask[:, :, top: top + patch_h, left: left + patch_w, front: front + patch_d] = 0.0
    loss_mask[:, top: top + patch_h, left: left + patch_w, front: front + patch_d] = 0.0

    return mask, loss_mask


# -------------------------
# Models / EMA
# -------------------------
class BCPNet(nn.Module):
    """
    Wrapper for student / teacher networks.
    """
    def __init__(self, model_name: str, in_chns: int, num_classes: int, ema: bool = False):
        super().__init__()
        model_name = model_name.lower()
        if model_name == "unet":
            self.net = UNet3D(in_channels=in_chns, num_classes=num_classes)
        elif model_name == "resunet":
            self.net = ResUNet3D(in_channels=in_chns, num_classes=num_classes)
        else:
            raise ValueError(f"Unknown model_name {model_name}")

        if ema:
            for p in self.net.parameters():
                p.detach_()

    def forward(self, x):
        return self.net(x)


def update_ema_variables(model, ema_model, alpha=0.99):
    """
    EMA update for teacher model.
    """
    for ema_param, param in zip(ema_model.parameters(), model.parameters()):
        ema_param.data.mul_(alpha).add_(param.data, alpha=1 - alpha)


# -------------------------
# Metrics: Dice + 95HD + ASD
# -------------------------
def dice_metric_from_logits(logits, targets, num_classes: int):
    """
    Compute mean foreground Dice for evaluation.
    logits: [B, C, H, W, D]
    targets: [B, H, W, D]
    """
    with torch.no_grad():
        probs = F.softmax(logits, dim=1)
        preds = torch.argmax(probs, dim=1)

        dices = []
        for cls in range(1, num_classes):  # ignore background = 0
            pred_c = (preds == cls).float()
            target_c = (targets == cls).float()
            inter = (pred_c * target_c).sum()
            denom = pred_c.sum() + target_c.sum()
            if denom > 0:
                dices.append((2.0 * inter + 1e-5) / (denom + 1e-5))
        if len(dices) == 0:
            return 0.0
        return float(torch.mean(torch.stack(dices)))


def _compute_binary_surface_distances(pred, gt, spacing=(1.0, 1.0, 1.0)):
    """
    pred, gt: boolean numpy arrays with shape [H, W, D]
    spacing: voxel spacing (z,y,x) or (H,W,D) order; here we assume (1,1,1).

    Returns:
      all_distances: np.array of surface distances (mm) between pred and gt surfaces.
    """
    pred = pred.astype(bool)
    gt = gt.astype(bool)

    if not pred.any() and not gt.any():
        return np.array([0.0])

    # surface = object - eroded_object
    pred_erode = binary_erosion(pred)
    gt_erode = binary_erosion(gt)
    pred_surface = pred ^ pred_erode
    gt_surface = gt ^ gt_erode

    if not pred_surface.any():
        pred_surface = pred  # fallback if no surface
    if not gt_surface.any():
        gt_surface = gt

    # distance transform of the complement
    dt_gt = distance_transform_edt(~gt, sampling=spacing)
    dt_pred = distance_transform_edt(~pred, sampling=spacing)

    # distances from pred surface to gt
    pred_surface_distances = dt_gt[pred_surface]
    # distances from gt surface to pred
    gt_surface_distances = dt_pred[gt_surface]

    all_distances = np.concatenate([pred_surface_distances, gt_surface_distances])
    if all_distances.size == 0:
        return np.array([0.0])

    return all_distances


def compute_95hd_asd(pred, gt, spacing=(1.0, 1.0, 1.0)):
    """
    Compute 95% Hausdorff Distance and Average Surface Distance
    for binary foreground masks.
    pred, gt: numpy arrays, shape [H,W,D] or [D,H,W], but assumed consistent.
    """
    if not pred.any() and not gt.any():
        # No foreground in both: perfect match
        return 0.0, 0.0
    if not pred.any() or not gt.any():
        # One empty, one not — can define as large distance or still compute;
        # we treat as large but finite, via surface distances
        pass

    dists = _compute_binary_surface_distances(pred, gt, spacing=spacing)
    hd95 = float(np.percentile(dists, 95))
    asd = float(np.mean(dists))
    return hd95, asd


def validate_model(model, val_loader, num_classes: int):
    """
    Run validation and return:
      - mean Dice
      - mean 95HD (foreground union over all classes > 0)
      - mean ASD

    For HD95/ASD we treat *any* non-background label as foreground.
    """
    model.eval()
    dices = []
    hd95s = []
    asds = []

    with torch.no_grad():
        for img, lab in val_loader:
            img = img.cuda()
            lab = lab.cuda().long().squeeze(1)  # [B, H, W, D]
            logits = model(img)

            # Dice (multi-class average over foreground classes)
            d = dice_metric_from_logits(logits, lab, num_classes)
            dices.append(d)

            # Compute HD95 & ASD on combined foreground (lab > 0)
            probs = F.softmax(logits, dim=1)
            preds = torch.argmax(probs, dim=1)  # [B, H, W, D]

            for b in range(preds.shape[0]):
                pred_b = preds[b].cpu().numpy()
                lab_b = lab[b].cpu().numpy()

                pred_fg = (pred_b > 0)
                gt_fg = (lab_b > 0)

                hd95, asd = compute_95hd_asd(pred_fg, gt_fg, spacing=(1.0, 1.0, 1.0))
                hd95s.append(hd95)
                asds.append(asd)

    if len(dices) == 0:
        return 0.0, 0.0, 0.0

    dice_mean = float(np.mean(dices))
    hd95_mean = float(np.mean(hd95s)) if len(hd95s) > 0 else 0.0
    asd_mean = float(np.mean(asds)) if len(asds) > 0 else 0.0
    return dice_mean, hd95_mean, asd_mean


# -------------------------
# Arg parser / Excel setup
# -------------------------
def get_parser():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--root_path",
        type=str,
        required=True,
        help="Root path of PiCAI data (train/ valid/ unlab_data/)",
    )
    parser.add_argument("--gpu", type=str, default="0")
    parser.add_argument("--batch_size", type=int, default=2)
    parser.add_argument("--max_epoch", type=int, default=300)
    parser.add_argument("--num_classes", type=int, default=2)
    parser.add_argument("--u_weight", type=float, default=0.5)
    parser.add_argument("--seed", type=int, default=1337)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--snapshot_path", type=str, default="./runs_sdcl_picai")

    # supervised pre-training epochs
    parser.add_argument(
        "--pretrain_epochs",
        type=int,
        default=20,
        help="Number of supervised warmup epochs before semi-supervised training",
    )

    return parser


def setup_excel(excel_path):
    """
    Create or load an Excel file with 'Warmup' and 'SSL' sheets.
    Each row:
      Epoch, Dice_S1, Dice_S2, Dice_EMA, Best_EMA_Dice,
      HD95_S1, HD95_S2, HD95_EMA, ASD_S1, ASD_S2, ASD_EMA
    """
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if "Warmup" not in wb.sheetnames:
            ws_warmup = wb.create_sheet("Warmup")
            ws_warmup.append([
                "Epoch",
                "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
                "HD95_Student1", "HD95_Student2", "HD95_EMA",
                "ASD_Student1", "ASD_Student2", "ASD_EMA",
            ])
        if "SSL" not in wb.sheetnames:
            ws_ssl = wb.create_sheet("SSL")
            ws_ssl.append([
                "Epoch",
                "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
                "HD95_Student1", "HD95_Student2", "HD95_EMA",
                "ASD_Student1", "ASD_Student2", "ASD_EMA",
            ])
    else:
        wb = Workbook()
        ws_warmup = wb.active
        ws_warmup.title = "Warmup"
        ws_warmup.append([
            "Epoch",
            "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
            "HD95_Student1", "HD95_Student2", "HD95_EMA",
            "ASD_Student1", "ASD_Student2", "ASD_EMA",
        ])
        ws_ssl = wb.create_sheet("SSL")
        ws_ssl.append([
            "Epoch",
            "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
            "HD95_Student1", "HD95_Student2", "HD95_EMA",
            "ASD_Student1", "ASD_Student2", "ASD_EMA",
        ])

    wb.save(excel_path)
    return wb


# -------------------------
# Supervised warmup
# -------------------------
def supervised_warmup(
    args,
    model,
    model2,
    ema_model,
    lab_loader_a,
    lab_loader_b,
    val_loader,
    criterion_sup,
    optimizer,
    optimizer2,
    excel_wb,
    excel_path,
    global_best_dict,
):
    """
    Supervised pre-training on labeled data only, as in SDCL paper.
    Students A (UNet) and B (ResUNet) are trained with DiceCE loss,
    and EMA teacher is updated from student A.
    Also does validation after each epoch and logs Dice/95HD/ASD to Excel.
    """
    logging.info(f"Starting supervised warmup for {args.pretrain_epochs} epochs")

    best_ema_dice = 0.0
    ws_warmup = excel_wb["Warmup"]

    for epoch in range(args.pretrain_epochs):
        model.train()
        model2.train()
        ema_model.train()

        loader_iter = zip(lab_loader_a, lab_loader_b)
        progress = tqdm(loader_iter, ncols=100, desc=f"Warmup {epoch + 1}/{args.pretrain_epochs}")

        for (img_a, lab_a), (img_b, lab_b) in progress:
            img_a = img_a.cuda()
            lab_a = lab_a.cuda().long().squeeze(1)
            img_b = img_b.cuda()
            lab_b = lab_b.cuda().long().squeeze(1)

            # ----- student 1 (UNet) -----
            optimizer.zero_grad()
            out_a_1 = model(img_a)
            out_b_1 = model(img_b)
            loss_sup_1 = (
                criterion_sup(out_a_1, lab_a) +
                criterion_sup(out_b_1, lab_b)
            ) / 2.0
            loss_sup_1.backward()
            optimizer.step()

            # ----- student 2 (ResUNet) -----
            optimizer2.zero_grad()
            out_a_2 = model2(img_a)
            out_b_2 = model2(img_b)
            loss_sup_2 = (
                criterion_sup(out_a_2, lab_a) +
                criterion_sup(out_b_2, lab_b)
            ) / 2.0
            loss_sup_2.backward()
            optimizer2.step()

            # ----- update EMA teacher from student A (UNet) -----
            update_ema_variables(model, ema_model, alpha=0.99)

            progress.set_postfix(
                {
                    "sup1": f"{loss_sup_1.item():.4f}",
                    "sup2": f"{loss_sup_2.item():.4f}",
                }
            )

        # --------- Validation after each warmup epoch ---------
        dice_s1, hd95_s1, asd_s1 = validate_model(model, val_loader, args.num_classes)
        dice_s2, hd95_s2, asd_s2 = validate_model(model2, val_loader, args.num_classes)
        dice_ema, hd95_ema, asd_ema = validate_model(ema_model, val_loader, args.num_classes)

        # Print Dice to console
        print(
            f"[Warmup {epoch + 1}/{args.pretrain_epochs}] "
            f"Dice_S1={dice_s1:.4f}, Dice_S2={dice_s2:.4f}, EMA={dice_ema:.4f}"
        )

        # phase-specific best tracking (just for logging)
        if dice_ema > best_ema_dice:
            best_ema_dice = dice_ema

        # global best model saving (over whole training)
        if dice_ema > global_best_dict["best_ema"]:
            global_best_dict["best_ema"] = dice_ema
            torch.save(model.state_dict(), os.path.join(args.snapshot_path, "best_studentA.pth"))
            torch.save(model2.state_dict(), os.path.join(args.snapshot_path, "best_studentB.pth"))
            torch.save(ema_model.state_dict(), os.path.join(args.snapshot_path, "best_ema.pth"))
            print(f"✔️ New GLOBAL BEST (Warmup epoch {epoch + 1}) saved with EMA Dice={dice_ema:.4f}")

        # log everything to Excel (Dice + 95HD + ASD)
        ws_warmup.append([
            epoch + 1,
            dice_s1, dice_s2, dice_ema, best_ema_dice,
            hd95_s1, hd95_s2, hd95_ema,
            asd_s1, asd_s2, asd_ema,
        ])
        excel_wb.save(excel_path)

        logging.info(
            f"[Warmup Epoch {epoch + 1}/{args.pretrain_epochs}] "
            f"Val Dice - Student1: {dice_s1:.4f}, "
            f"Student2: {dice_s2:.4f}, EMA: {dice_ema:.4f}, "
            f"Best EMA (warmup phase): {best_ema_dice:.4f}"
        )

    logging.info("Warmup finished.")


# -------------------------
# Main training (SSL)
# -------------------------
def train(args):
    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

    snapshot_path = args.snapshot_path
    os.makedirs(snapshot_path, exist_ok=True)

    # logging
    logging.basicConfig(
        filename=os.path.join(snapshot_path, "log.txt"),
        level=logging.INFO,
        format="[%(asctime)s.%(msecs)03d] %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.getLogger().addHandler(logging.StreamHandler())
    logging.info(str(args))

    excel_path = os.path.join(snapshot_path, "metrics.xlsx")
    print("Saving metrics to:", os.path.abspath(excel_path))
    excel_wb = setup_excel(excel_path)

    # reproducibility
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    torch.cuda.manual_seed_all(args.seed)
    cudnn.benchmark = False

    # dataloaders
    (
        lab_loader_a,
        lab_loader_b,
        unlab_loader_a,
        unlab_loader_b,
        val_loader,
    ) = get_sdcl_dataloaders(
        args.root_path,
        batch_size=args.batch_size,
        num_workers=4,
        seed=args.seed,
    )

    num_classes = args.num_classes
    in_chns = 3

    # models
    model = BCPNet(model_name="unet", in_chns=in_chns, num_classes=num_classes).cuda()
    model2 = BCPNet(model_name="resunet", in_chns=in_chns, num_classes=num_classes).cuda()
    ema_model = BCPNet(model_name="unet", in_chns=in_chns, num_classes=num_classes, ema=True).cuda()

    optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)
    optimizer2 = torch.optim.Adam(model2.parameters(), lr=args.lr)

    criterion_sup = DiceCELoss(num_classes=num_classes)

    # initialize EMA from student A before warmup
    ema_model.load_state_dict(model.state_dict())

    # global best tracker over whole training (warmup + SSL)
    global_best = {"best_ema": 0.0}

    # --------------------
    # supervised warmup phase
    # --------------------
    if args.pretrain_epochs > 0:
        supervised_warmup(
            args,
            model,
            model2,
            ema_model,
            lab_loader_a,
            lab_loader_b,
            val_loader,
            criterion_sup,
            optimizer,
            optimizer2,
            excel_wb,
            excel_path,
            global_best,
        )
    else:
        logging.info("Skipping supervised warmup (pretrain_epochs == 0).")

    # reload excel workbook for SSL logging
    excel_wb = load_workbook(excel_path)
    ws_ssl = excel_wb["SSL"]
    best_ssl_ema_dice = 0.0

    # --------------------
    # semi-supervised SDCL + BCP phase
    # --------------------
    max_epoch = args.max_epoch
    global_step = 0

    for epoch in range(max_epoch):
        model.train()
        model2.train()
        ema_model.train()

        loader_iter = zip(lab_loader_a, lab_loader_b, unlab_loader_a, unlab_loader_b)
        progress = tqdm(loader_iter, ncols=100, desc=f"SSL Epoch {epoch + 1}/{max_epoch}")

        for (img_a, lab_a), (img_b, lab_b), (unimg_a, _), (unimg_b, _) in progress:
            img_a = img_a.cuda()
            lab_a = lab_a.cuda().long().squeeze(1)  # [B, H, W, D]
            img_b = img_b.cuda()
            lab_b = lab_b.cuda().long().squeeze(1)
            unimg_a = unimg_a.cuda()
            unimg_b = unimg_b.cuda()

            # teacher pseudo labels for unlabeled images
            with torch.no_grad():
                pre_a = ema_model(unimg_a)
                pre_b = ema_model(unimg_b)

                plab_a = torch.argmax(F.softmax(pre_a, dim=1), dim=1)
                plab_b = torch.argmax(F.softmax(pre_b, dim=1), dim=1)

                img_mask, loss_mask = generate_3d_mask(img_a)

            # mix labeled & unlabeled (BCP style)
            net_input_l = unimg_a * img_mask + img_b * (1.0 - img_mask)
            net_input_unl = img_a * img_mask + unimg_b * (1.0 - img_mask)

            out_l = model(net_input_l)
            out_unl = model(net_input_unl)

            out_l_2 = model2(net_input_l)
            out_unl_2 = model2(net_input_unl)

            # SDCL mix losses for net1
            l_dice, l_ce = mix_loss(
                out_l,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
            )
            unl_dice, unl_ce = mix_loss(
                out_unl,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
            )

            # SDCL mix losses for net2
            l_dice_2, l_ce_2 = mix_loss(
                out_l_2,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
            )
            unl_dice_2, unl_ce_2 = mix_loss(
                out_unl_2,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
            )

            loss_ce = unl_ce + l_ce
            loss_dice = unl_dice + l_dice

            loss_ce_2 = unl_ce_2 + l_ce_2
            loss_dice_2 = unl_dice_2 + l_dice_2

            with torch.no_grad():
                diff_mask1 = get_XOR_region(out_l, out_l_2)
                diff_mask2 = get_XOR_region(out_unl, out_unl_2)

            # discrepancy-aware MSE & KL (net1)
            net1_mse_loss_lab = mix_mse_loss(
                out_l,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net1_mse_loss_unlab = mix_mse_loss(
                out_unl,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )
            net1_kl_loss_lab = mix_max_kl_loss(
                out_l,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net1_kl_loss_unlab = mix_max_kl_loss(
                out_unl,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )

            # discrepancy-aware MSE & KL (net2)
            net2_mse_loss_lab = mix_mse_loss(
                out_l_2,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net2_mse_loss_unlab = mix_mse_loss(
                out_unl_2,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )
            net2_kl_loss_lab = mix_max_kl_loss(
                out_l_2,
                plab_a,
                lab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=True,
                diff_mask=diff_mask1,
            )
            net2_kl_loss_unlab = mix_max_kl_loss(
                out_unl_2,
                lab_a,
                plab_b,
                loss_mask,
                num_classes=num_classes,
                l_weight=1.0,
                u_weight=args.u_weight,
                unlab=False,
                diff_mask=diff_mask2,
            )

            # loss weights (γ = 0.5, μ = 0.05)
            loss1 = (
                (loss_dice + loss_ce) / 2.0
                + 0.5 * (net1_mse_loss_lab + net1_mse_loss_unlab)
                + 0.05 * (net1_kl_loss_lab + net1_kl_loss_unlab)
            )
            loss2 = (
                (loss_dice_2 + loss_ce_2) / 2.0
                + 0.5 * (net2_mse_loss_lab + net2_mse_loss_unlab)
                + 0.05 * (net2_kl_loss_lab + net2_kl_loss_unlab)
            )

            optimizer.zero_grad()
            loss1.backward()
            optimizer.step()

            optimizer2.zero_grad()
            loss2.backward()
            optimizer2.step()

            update_ema_variables(model, ema_model, alpha=0.99)

            global_step += 1
            progress.set_postfix(
                {
                    "loss1": f"{loss1.item():.4f}",
                    "loss2": f"{loss2.item():.4f}",
                }
            )

        # --------- Validation after each SSL epoch ---------
        dice_s1, hd95_s1, asd_s1 = validate_model(model, val_loader, args.num_classes)
        dice_s2, hd95_s2, asd_s2 = validate_model(model2, val_loader, args.num_classes)
        dice_ema, hd95_ema, asd_ema = validate_model(ema_model, val_loader, args.num_classes)

        # Print Dice to console
        print(
            f"[SSL {epoch + 1}/{max_epoch}] "
            f"Dice_S1={dice_s1:.4f}, Dice_S2={dice_s2:.4f}, EMA={dice_ema:.4f}"
        )

        # phase-specific best tracking (for logging only)
        if dice_ema > best_ssl_ema_dice:
            best_ssl_ema_dice = dice_ema

        # global best model saving (over whole training)
        if dice_ema > global_best["best_ema"]:
            global_best["best_ema"] = dice_ema
            torch.save(model.state_dict(), os.path.join(snapshot_path, "best_studentA.pth"))
            torch.save(model2.state_dict(), os.path.join(snapshot_path, "best_studentB.pth"))
            torch.save(ema_model.state_dict(), os.path.join(snapshot_path, "best_ema.pth"))
            print(f" New GLOBAL BEST (SSL epoch {epoch + 1}) saved with EMA Dice={dice_ema:.4f}")

        # log to Excel (Dice + 95HD + ASD)
        ws_ssl.append([
            epoch + 1,
            dice_s1, dice_s2, dice_ema, best_ssl_ema_dice,
            hd95_s1, hd95_s2, hd95_ema,
            asd_s1, asd_s2, asd_ema,
        ])
        excel_wb.save(excel_path)

        logging.info(
            f"[SSL Epoch {epoch + 1}/{max_epoch}] "
            f"Val Dice - Student1: {dice_s1:.4f}, "
            f"Student2: {dice_s2:.4f}, EMA: {dice_ema:.4f}, "
            f"Best SSL EMA (phase): {best_ssl_ema_dice:.4f}"
        )

    logging.info("Training finished.")


if __name__ == "__main__":
    parser = get_parser()
    args = parser.parse_args()
    train(args)



# import os
# import argparse
# import logging
# from tqdm import tqdm

# import numpy as np
# import torch
# import torch.nn as nn
# import torch.nn.functional as F
# import torch.backends.cudnn as cudnn

# from openpyxl import Workbook, load_workbook
# from scipy.ndimage import distance_transform_edt, binary_erosion

# from picai_dataset import get_sdcl_dataloaders
# from models.unet3d import UNet3D
# from models.resunet3d import ResUNet3D
# from losses import (
#     DiceCELoss,
#     mix_loss,
#     mix_mse_loss,
#     mix_max_kl_loss,
#     get_XOR_region,
# )


# # -------------------------
# # Utility: BCP mask
# # -------------------------
# def generate_3d_mask(img):
#     """
#     img: [B, C, H, W, D]
#     Returns:
#       img_mask:  [B, 1, H, W, D]  (for mixing input images)
#       loss_mask: [B, H, W, D]     (for region-weighted losses)
#     """
#     b, c, h, w, d = img.shape
#     device = img.device

#     mask = torch.ones((b, 1, h, w, d), device=device)
#     loss_mask = torch.ones((b, h, w, d), device=device)

#     patch_h = int(h * 2 / 3)
#     patch_w = int(w * 2 / 3)
#     patch_d = int(max(1, d * 2 / 3))

#     top = np.random.randint(0, max(1, h - patch_h + 1))
#     left = np.random.randint(0, max(1, w - patch_w + 1))
#     front = np.random.randint(0, max(1, d - patch_d + 1))

#     mask[:, :, top: top + patch_h, left: left + patch_w, front: front + patch_d] = 0.0
#     loss_mask[:, top: top + patch_h, left: left + patch_w, front: front + patch_d] = 0.0

#     return mask, loss_mask


# # -------------------------
# # Models / EMA
# # -------------------------
# class BCPNet(nn.Module):
#     """
#     Wrapper for student / teacher networks.
#     """
#     def __init__(self, model_name: str, in_chns: int, num_classes: int, ema: bool = False):
#         super().__init__()
#         model_name = model_name.lower()
#         if model_name == "unet":
#             self.net = UNet3D(in_channels=in_chns, num_classes=num_classes)
#         elif model_name == "resunet":
#             self.net = ResUNet3D(in_channels=in_chns, num_classes=num_classes)
#         else:
#             raise ValueError(f"Unknown model_name {model_name}")

#         if ema:
#             for p in self.net.parameters():
#                 p.detach_()

#     def forward(self, x):
#         return self.net(x)


# def update_ema_variables(model, ema_model, alpha=0.99):
#     """
#     EMA update for teacher model.
#     """
#     for ema_param, param in zip(ema_model.parameters(), model.parameters()):
#         ema_param.data.mul_(alpha).add_(param.data, alpha=1 - alpha)


# # -------------------------
# # Metrics: Dice + 95HD + ASD
# # -------------------------
# def dice_metric_from_logits(logits, targets, num_classes: int):
#     """
#     Compute mean foreground Dice for evaluation.
#     logits: [B, C, H, W, D]
#     targets: [B, H, W, D]
#     """
#     with torch.no_grad():
#         probs = F.softmax(logits, dim=1)
#         preds = torch.argmax(probs, dim=1)

#         dices = []
#         for cls in range(1, num_classes):  # ignore background = 0
#             pred_c = (preds == cls).float()
#             target_c = (targets == cls).float()
#             inter = (pred_c * target_c).sum()
#             denom = pred_c.sum() + target_c.sum()
#             if denom > 0:
#                 dices.append((2.0 * inter + 1e-5) / (denom + 1e-5))
#         if len(dices) == 0:
#             return 0.0
#         return float(torch.mean(torch.stack(dices)))


# def _compute_binary_surface_distances(pred, gt, spacing=(1.0, 1.0, 1.0)):
#     """
#     pred, gt: boolean numpy arrays with shape [H, W, D]
#     spacing: voxel spacing (z,y,x) or (H,W,D) order; here we assume (1,1,1).

#     Returns:
#       all_distances: np.array of surface distances (mm) between pred and gt surfaces.
#     """
#     pred = pred.astype(bool)
#     gt = gt.astype(bool)

#     if not pred.any() and not gt.any():
#         return np.array([0.0])

#     # surface = object - eroded_object
#     pred_erode = binary_erosion(pred)
#     gt_erode = binary_erosion(gt)
#     pred_surface = pred ^ pred_erode
#     gt_surface = gt ^ gt_erode

#     if not pred_surface.any():
#         pred_surface = pred  # fallback if no surface
#     if not gt_surface.any():
#         gt_surface = gt

#     # distance transform of the complement
#     dt_gt = distance_transform_edt(~gt, sampling=spacing)
#     dt_pred = distance_transform_edt(~pred, sampling=spacing)

#     # distances from pred surface to gt
#     pred_surface_distances = dt_gt[pred_surface]
#     # distances from gt surface to pred
#     gt_surface_distances = dt_pred[gt_surface]

#     all_distances = np.concatenate([pred_surface_distances, gt_surface_distances])
#     if all_distances.size == 0:
#         return np.array([0.0])

#     return all_distances


# def compute_95hd_asd(pred, gt, spacing=(1.0, 1.0, 1.0)):
#     """
#     Compute 95% Hausdorff Distance and Average Surface Distance
#     for binary foreground masks.
#     pred, gt: numpy arrays, shape [H,W,D] or [D,H,W], but assumed consistent.
#     """
#     if not pred.any() and not gt.any():
#         # No foreground in both: perfect match
#         return 0.0, 0.0
#     if not pred.any() or not gt.any():
#         # One empty, one not — can define as large distance or still compute;
#         # we treat as large but finite, via surface distances
#         # (the DT-based function will handle it).
#         pass

#     dists = _compute_binary_surface_distances(pred, gt, spacing=spacing)
#     hd95 = float(np.percentile(dists, 95))
#     asd = float(np.mean(dists))
#     return hd95, asd


# def validate_model(model, val_loader, num_classes: int):
#     """
#     Run validation and return:
#       - mean Dice
#       - mean 95HD (foreground union over all classes > 0)
#       - mean ASD

#     For HD95/ASD we treat *any* non-background label as foreground.
#     """
#     model.eval()
#     dices = []
#     hd95s = []
#     asds = []

#     with torch.no_grad():
#         for img, lab in val_loader:
#             img = img.cuda()
#             lab = lab.cuda().long().squeeze(1)  # [B, H, W, D]
#             logits = model(img)

#             # Dice (multi-class average over foreground classes)
#             d = dice_metric_from_logits(logits, lab, num_classes)
#             dices.append(d)

#             # Compute HD95 & ASD on combined foreground (lab > 0)
#             probs = F.softmax(logits, dim=1)
#             preds = torch.argmax(probs, dim=1)  # [B, H, W, D]

#             for b in range(preds.shape[0]):
#                 pred_b = preds[b].cpu().numpy()
#                 lab_b = lab[b].cpu().numpy()

#                 pred_fg = (pred_b > 0)
#                 gt_fg = (lab_b > 0)

#                 hd95, asd = compute_95hd_asd(pred_fg, gt_fg, spacing=(1.0, 1.0, 1.0))
#                 hd95s.append(hd95)
#                 asds.append(asd)

#     if len(dices) == 0:
#         return 0.0, 0.0, 0.0

#     dice_mean = float(np.mean(dices))
#     hd95_mean = float(np.mean(hd95s)) if len(hd95s) > 0 else 0.0
#     asd_mean = float(np.mean(asds)) if len(asds) > 0 else 0.0
#     return dice_mean, hd95_mean, asd_mean


# # -------------------------
# # Arg parser / Excel setup
# # -------------------------
# def get_parser():
#     parser = argparse.ArgumentParser()

#     parser.add_argument(
#         "--root_path",
#         type=str,
#         required=True,
#         help="Root path of PiCAI data (train/ valid/ unlab_data/)",
#     )
#     parser.add_argument("--gpu", type=str, default="0")
#     parser.add_argument("--batch_size", type=int, default=2)
#     parser.add_argument("--max_epoch", type=int, default=300)
#     parser.add_argument("--num_classes", type=int, default=2)
#     parser.add_argument("--u_weight", type=float, default=0.5)
#     parser.add_argument("--seed", type=int, default=1337)
#     parser.add_argument("--lr", type=float, default=1e-3)
#     parser.add_argument("--snapshot_path", type=str, default="./runs_sdcl_picai")

#     # supervised pre-training epochs
#     parser.add_argument(
#         "--pretrain_epochs",
#         type=int,
#         default=20,
#         help="Number of supervised warmup epochs before semi-supervised training",
#     )

#     return parser


# def setup_excel(excel_path):
#     """
#     Create or load an Excel file with 'Warmup' and 'SSL' sheets.
#     Each row:
#       Epoch, Dice_S1, Dice_S2, Dice_EMA, Best_EMA_Dice,
#       HD95_S1, HD95_S2, HD95_EMA, ASD_S1, ASD_S2, ASD_EMA
#     """
#     if os.path.exists(excel_path):
#         wb = load_workbook(excel_path)
#         if "Warmup" not in wb.sheetnames:
#             ws_warmup = wb.create_sheet("Warmup")
#             ws_warmup.append([
#                 "Epoch",
#                 "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
#                 "HD95_Student1", "HD95_Student2", "HD95_EMA",
#                 "ASD_Student1", "ASD_Student2", "ASD_EMA",
#             ])
#         if "SSL" not in wb.sheetnames:
#             ws_ssl = wb.create_sheet("SSL")
#             ws_ssl.append([
#                 "Epoch",
#                 "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
#                 "HD95_Student1", "HD95_Student2", "HD95_EMA",
#                 "ASD_Student1", "ASD_Student2", "ASD_EMA",
#             ])
#     else:
#         wb = Workbook()
#         ws_warmup = wb.active
#         ws_warmup.title = "Warmup"
#         ws_warmup.append([
#             "Epoch",
#             "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
#             "HD95_Student1", "HD95_Student2", "HD95_EMA",
#             "ASD_Student1", "ASD_Student2", "ASD_EMA",
#         ])
#         ws_ssl = wb.create_sheet("SSL")
#         ws_ssl.append([
#             "Epoch",
#             "Dice_Student1", "Dice_Student2", "Dice_EMA", "Best_EMA_Dice",
#             "HD95_Student1", "HD95_Student2", "HD95_EMA",
#             "ASD_Student1", "ASD_Student2", "ASD_EMA",
#         ])

#     wb.save(excel_path)
#     return wb


# # -------------------------
# # Supervised warmup
# # -------------------------
# def supervised_warmup(
#     args,
#     model,
#     model2,
#     ema_model,
#     lab_loader_a,
#     lab_loader_b,
#     val_loader,
#     criterion_sup,
#     optimizer,
#     optimizer2,
#     excel_wb,
#     excel_path,
# ):
#     """
#     Supervised pre-training on labeled data only, as in SDCL paper.
#     Students A (UNet) and B (ResUNet) are trained with DiceCE loss,
#     and EMA teacher is updated from student A.
#     Also does validation after each epoch and logs Dice/95HD/ASD to Excel.
#     """
#     logging.info(f"Starting supervised warmup for {args.pretrain_epochs} epochs")

#     best_ema_dice = 0.0
#     ws_warmup = excel_wb["Warmup"]

#     for epoch in range(args.pretrain_epochs):
#         model.train()
#         model2.train()
#         ema_model.train()

#         loader_iter = zip(lab_loader_a, lab_loader_b)
#         progress = tqdm(loader_iter, ncols=100, desc=f"Warmup {epoch + 1}/{args.pretrain_epochs}")

#         for (img_a, lab_a), (img_b, lab_b) in progress:
#             img_a = img_a.cuda()
#             lab_a = lab_a.cuda().long().squeeze(1)
#             img_b = img_b.cuda()
#             lab_b = lab_b.cuda().long().squeeze(1)

#             # ----- student 1 (UNet) -----
#             optimizer.zero_grad()
#             out_a_1 = model(img_a)
#             out_b_1 = model(img_b)
#             loss_sup_1 = (
#                 criterion_sup(out_a_1, lab_a) +
#                 criterion_sup(out_b_1, lab_b)
#             ) / 2.0
#             loss_sup_1.backward()
#             optimizer.step()

#             # ----- student 2 (ResUNet) -----
#             optimizer2.zero_grad()
#             out_a_2 = model2(img_a)
#             out_b_2 = model2(img_b)
#             loss_sup_2 = (
#                 criterion_sup(out_a_2, lab_a) +
#                 criterion_sup(out_b_2, lab_b)
#             ) / 2.0
#             loss_sup_2.backward()
#             optimizer2.step()

#             # ----- update EMA teacher from student A (UNet) -----
#             update_ema_variables(model, ema_model, alpha=0.99)

#             progress.set_postfix(
#                 {
#                     "sup1": f"{loss_sup_1.item():.4f}",
#                     "sup2": f"{loss_sup_2.item():.4f}",
#                 }
#             )

#         # --------- Validation after each warmup epoch ---------
#         dice_s1, hd95_s1, asd_s1 = validate_model(model, val_loader, args.num_classes)
#         dice_s2, hd95_s2, asd_s2 = validate_model(model2, val_loader, args.num_classes)
#         dice_ema, hd95_ema, asd_ema = validate_model(ema_model, val_loader, args.num_classes)

#         # Print Dice to console
#         print(
#             f"[Warmup {epoch + 1}/{args.pretrain_epochs}] "
#             f"Dice_S1={dice_s1:.4f}, Dice_S2={dice_s2:.4f}, EMA={dice_ema:.4f}"
#         )

#         if dice_ema > best_ema_dice:
#             best_ema_dice = dice_ema
#             # save best warmup models
#             torch.save(model.state_dict(), os.path.join(args.snapshot_path, "best_warmup_studentA.pth"))
#             torch.save(model2.state_dict(), os.path.join(args.snapshot_path, "best_warmup_studentB.pth"))
#             torch.save(ema_model.state_dict(), os.path.join(args.snapshot_path, "best_warmup_ema.pth"))

#         # log everything to Excel (Dice + 95HD + ASD)
#         ws_warmup.append([
#             epoch + 1,
#             dice_s1, dice_s2, dice_ema, best_ema_dice,
#             hd95_s1, hd95_s2, hd95_ema,
#             asd_s1, asd_s2, asd_ema,
#         ])
#         excel_wb.save(excel_path)

#         # print ONLY Dice to command line via logging
#         logging.info(
#             f"[Warmup Epoch {epoch + 1}/{args.pretrain_epochs}] "
#             f"Val Dice - Student1: {dice_s1:.4f}, "
#             f"Student2: {dice_s2:.4f}, EMA: {dice_ema:.4f}, "
#             f"Best EMA: {best_ema_dice:.4f}"
#         )

#     logging.info("Warmup finished.")


# # -------------------------
# # Main training (SSL)
# # -------------------------
# def train(args):
#     os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

#     snapshot_path = args.snapshot_path
#     os.makedirs(snapshot_path, exist_ok=True)

#     # logging
#     logging.basicConfig(
#         filename=os.path.join(snapshot_path, "log.txt"),
#         level=logging.INFO,
#         format="[%(asctime)s.%(msecs)03d] %(message)s",
#         datefmt="%H:%M:%S",
#     )
#     logging.getLogger().addHandler(logging.StreamHandler())
#     logging.info(str(args))

#     excel_path = os.path.join(snapshot_path, "metrics.xlsx")
#     print("Saving metrics to:", os.path.abspath(excel_path))  # to know exactly which file
#     excel_wb = setup_excel(excel_path)

#     # reproducibility
#     np.random.seed(args.seed)
#     torch.manual_seed(args.seed)
#     torch.cuda.manual_seed_all(args.seed)
#     cudnn.benchmark = False

#     # dataloaders
#     (
#         lab_loader_a,
#         lab_loader_b,
#         unlab_loader_a,
#         unlab_loader_b,
#         val_loader,
#     ) = get_sdcl_dataloaders(
#         args.root_path,
#         batch_size=args.batch_size,
#         num_workers=4,
#         seed=args.seed,
#     )

#     num_classes = args.num_classes
#     in_chns = 3

#     # models
#     model = BCPNet(model_name="unet", in_chns=in_chns, num_classes=num_classes).cuda()
#     model2 = BCPNet(model_name="resunet", in_chns=in_chns, num_classes=num_classes).cuda()
#     ema_model = BCPNet(model_name="unet", in_chns=in_chns, num_classes=num_classes, ema=True).cuda()

#     optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)
#     optimizer2 = torch.optim.Adam(model2.parameters(), lr=args.lr)

#     criterion_sup = DiceCELoss(num_classes=num_classes)

#     # initialize EMA from student A before warmup
#     ema_model.load_state_dict(model.state_dict())

#     # --------------------
#     # supervised warmup phase
#     # --------------------
#     if args.pretrain_epochs > 0:
#         supervised_warmup(
#             args,
#             model,
#             model2,
#             ema_model,
#             lab_loader_a,
#             lab_loader_b,
#             val_loader,
#             criterion_sup,
#             optimizer,
#             optimizer2,
#             excel_wb,
#             excel_path,
#         )
#     else:
#         logging.info("Skipping supervised warmup (pretrain_epochs == 0).")

#     # reload excel workbook for SSL logging
#     excel_wb = load_workbook(excel_path)
#     ws_ssl = excel_wb["SSL"]
#     best_ssl_ema_dice = 0.0

#     # --------------------
#     # semi-supervised SDCL + BCP phase
#     # --------------------
#     max_epoch = args.max_epoch
#     global_step = 0

#     for epoch in range(max_epoch):
#         model.train()
#         model2.train()
#         ema_model.train()

#         loader_iter = zip(lab_loader_a, lab_loader_b, unlab_loader_a, unlab_loader_b)
#         progress = tqdm(loader_iter, ncols=100, desc=f"SSL Epoch {epoch + 1}/{max_epoch}")

#         for (img_a, lab_a), (img_b, lab_b), (unimg_a, _), (unimg_b, _) in progress:
#             img_a = img_a.cuda()
#             lab_a = lab_a.cuda().long().squeeze(1)  # [B, H, W, D]
#             img_b = img_b.cuda()
#             lab_b = lab_b.cuda().long().squeeze(1)
#             unimg_a = unimg_a.cuda()
#             unimg_b = unimg_b.cuda()

#             # teacher pseudo labels for unlabeled images
#             with torch.no_grad():
#                 pre_a = ema_model(unimg_a)
#                 pre_b = ema_model(unimg_b)

#                 plab_a = torch.argmax(F.softmax(pre_a, dim=1), dim=1)
#                 plab_b = torch.argmax(F.softmax(pre_b, dim=1), dim=1)

#                 img_mask, loss_mask = generate_3d_mask(img_a)

#             # mix labeled & unlabeled (BCP style)
#             net_input_l = unimg_a * img_mask + img_b * (1.0 - img_mask)
#             net_input_unl = img_a * img_mask + unimg_b * (1.0 - img_mask)

#             out_l = model(net_input_l)
#             out_unl = model(net_input_unl)

#             out_l_2 = model2(net_input_l)
#             out_unl_2 = model2(net_input_unl)

#             # SDCL mix losses for net1
#             l_dice, l_ce = mix_loss(
#                 out_l,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#             )
#             unl_dice, unl_ce = mix_loss(
#                 out_unl,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#             )

#             # SDCL mix losses for net2
#             l_dice_2, l_ce_2 = mix_loss(
#                 out_l_2,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#             )
#             unl_dice_2, unl_ce_2 = mix_loss(
#                 out_unl_2,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#             )

#             loss_ce = unl_ce + l_ce
#             loss_dice = unl_dice + l_dice

#             loss_ce_2 = unl_ce_2 + l_ce_2
#             loss_dice_2 = unl_dice_2 + l_dice_2

#             with torch.no_grad():
#                 diff_mask1 = get_XOR_region(out_l, out_l_2)
#                 diff_mask2 = get_XOR_region(out_unl, out_unl_2)

#             # discrepancy-aware MSE & KL (net1)
#             net1_mse_loss_lab = mix_mse_loss(
#                 out_l,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net1_mse_loss_unlab = mix_mse_loss(
#                 out_unl,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )
#             net1_kl_loss_lab = mix_max_kl_loss(
#                 out_l,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net1_kl_loss_unlab = mix_max_kl_loss(
#                 out_unl,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )

#             # discrepancy-aware MSE & KL (net2)
#             net2_mse_loss_lab = mix_mse_loss(
#                 out_l_2,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net2_mse_loss_unlab = mix_mse_loss(
#                 out_unl_2,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )
#             net2_kl_loss_lab = mix_max_kl_loss(
#                 out_l_2,
#                 plab_a,
#                 lab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=True,
#                 diff_mask=diff_mask1,
#             )
#             net2_kl_loss_unlab = mix_max_kl_loss(
#                 out_unl_2,
#                 lab_a,
#                 plab_b,
#                 loss_mask,
#                 num_classes=num_classes,
#                 l_weight=1.0,
#                 u_weight=args.u_weight,
#                 unlab=False,
#                 diff_mask=diff_mask2,
#             )

#             # loss weights (γ = 0.5, μ = 0.05)
#             loss1 = (
#                 (loss_dice + loss_ce) / 2.0
#                 + 0.5 * (net1_mse_loss_lab + net1_mse_loss_unlab)
#                 + 0.05 * (net1_kl_loss_lab + net1_kl_loss_unlab)
#             )
#             loss2 = (
#                 (loss_dice_2 + loss_ce_2) / 2.0
#                 + 0.5 * (net2_mse_loss_lab + net2_mse_loss_unlab)
#                 + 0.05 * (net2_kl_loss_lab + net2_kl_loss_unlab)
#             )

#             optimizer.zero_grad()
#             loss1.backward()
#             optimizer.step()

#             optimizer2.zero_grad()
#             loss2.backward()
#             optimizer2.step()

#             update_ema_variables(model, ema_model, alpha=0.99)

#             global_step += 1
#             progress.set_postfix(
#                 {
#                     "loss1": f"{loss1.item():.4f}",
#                     "loss2": f"{loss2.item():.4f}",
#                 }
#             )

#         # --------- Validation after each SSL epoch ---------
#         dice_s1, hd95_s1, asd_s1 = validate_model(model, val_loader, args.num_classes)
#         dice_s2, hd95_s2, asd_s2 = validate_model(model2, val_loader, args.num_classes)
#         dice_ema, hd95_ema, asd_ema = validate_model(ema_model, val_loader, args.num_classes)

#         # Print Dice to console
#         print(
#             f"[SSL {epoch + 1}/{max_epoch}] "
#             f"Dice_S1={dice_s1:.4f}, Dice_S2={dice_s2:.4f}, EMA={dice_ema:.4f}"
#         )

#         if dice_ema > best_ssl_ema_dice:
#             best_ssl_ema_dice = dice_ema
#             # save best SSL models
#             torch.save(model.state_dict(), os.path.join(snapshot_path, "best_ssl_studentA.pth"))
#             torch.save(model2.state_dict(), os.path.join(snapshot_path, "best_ssl_studentB.pth"))
#             torch.save(ema_model.state_dict(), os.path.join(snapshot_path, "best_ssl_ema.pth"))

#         # log to Excel (Dice + 95HD + ASD)
#         ws_ssl.append([
#             epoch + 1,
#             dice_s1, dice_s2, dice_ema, best_ssl_ema_dice,
#             hd95_s1, hd95_s2, hd95_ema,
#             asd_s1, asd_s2, asd_ema,
#         ])
#         excel_wb.save(excel_path)

#         # print ONLY Dice to command line via logging
#         logging.info(
#             f"[SSL Epoch {epoch + 1}/{max_epoch}] "
#             f"Val Dice - Student1: {dice_s1:.4f}, "
#             f"Student2: {dice_s2:.4f}, EMA: {dice_ema:.4f}, "
#             f"Best SSL EMA: {best_ssl_ema_dice:.4f}"
#         )

#     logging.info("Training finished.")


# if __name__ == "__main__":
#     parser = get_parser()
#     args = parser.parse_args()
#     train(args)
